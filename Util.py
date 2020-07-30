import openpyxl

import Logic
class Utils:
    def __init__(self):
        self.ext_txt = ".txt"
        self.ext_dat = ".dat"
        self.ext_xlsx = ".xlsx"
        self.deli = "^"
        self.indel_posit = 'O'
        self.indel_nega = 'X'
        self.percent = 100

    def csv_to_list_ignr_header(self, path, deli_str=","):
        result_list = []
        with open(path, "r") as f:
            header = f.readline()
            print(header)
            while True:
                tmp_line = f.readline().replace("\n", "")
                if tmp_line == '':
                    break

                result_list.append(tmp_line.split(deli_str))
        return result_list


    def make_excel_indel_frequency_by_cell_id(self, path, data_dict, names):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        total_data = [
            ['', names[0] + '_cnt', '', names[1] + '_cnt']
            , ['mut', data_dict['main_mut_cnt'][0], 'mut', data_dict['main_mut_cnt'][1]]
            , ['', '', 'non_mut', data_dict['main_mut_cnt'][2]]
            , ['non_mut', data_dict['main_non_mut_cnt'][0], 'mut', data_dict['main_non_mut_cnt'][1]]
            , ['', '', 'non_mut', data_dict['main_non_mut_cnt'][2]]
        ]

        for idx in range(len(total_data)):
            self.make_row(sheet, idx + 1, total_data[idx])

        workbook.save(filename=path + self.ext_xlsx)

    def make_excel_err_list(self, path, data_list):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        row = 1
        for data_arr in data_list[::-1]:
            col = 1
            for data_val in data_arr:
                sheet.cell(row=row, column=col, value=str(data_val))
                col += 1
            row += 1

        workbook.save(filename=path + self.ext_xlsx)

    def make_excel_homo_hetero(self, path, trgt_list, id_list):
        logic = Logic.Logics()

        workbook = openpyxl.Workbook()
        sheet = workbook.active

        main_mut_dict = trgt_list[0][0]
        main_non_mut_dict = trgt_list[0][1]
        sub_mut_dict = trgt_list[1][0]
        sub_non_mut_dict = trgt_list[1][1]

        row = 1
        cnt_sam = ['homo', 'hetero', 'wt']
        cnt_arr = []
        for h_1 in cnt_sam:
            for h_2 in cnt_sam:
                cnt_arr.append(h_1 + "_" + h_2)
        self.make_row(sheet, row, cnt_arr)

        row = 4
        header_arr = ['index', 'barcode_1', 'barcode_2', 'homo_hetero', 'main_NGS_read', 'indel_flag', '#of_read',
                      'sub_homo_hetero', 'sub_NGS_read', 'sub_indel_flag', 'sub_#of_read']
        self.make_row(sheet, row, header_arr)

        cell_non_junk_list = []
        non_cell_junk_list = []
        cnt_homo_hetero_wt = [0, 0, 0, 0, 0, 0, 0, 0, 0]
        for cell_id in id_list:
            index = 1

            # filter out cell_none or none_cell junks
            main_homo_hetero = logic.check_homo_hetero(cell_id, main_mut_dict, main_non_mut_dict)
            sub_homo_hetero = logic.check_homo_hetero(cell_id, sub_mut_dict, sub_non_mut_dict)
            if main_homo_hetero == "":
                if sub_homo_hetero == "":
                    continue
                else:  # filter out cell_none junks
                    logic.add_junk_list(cell_id, sub_mut_dict, sub_non_mut_dict, non_cell_junk_list)
                    continue
            # filter out none_cell junks
            if sub_homo_hetero == "":
                logic.add_junk_list(cell_id, main_mut_dict, main_non_mut_dict, cell_non_junk_list)
                continue

            logic.count_homo_hetero_wt(main_homo_hetero, sub_homo_hetero, cnt_homo_hetero_wt)

            if cell_id in main_mut_dict:
                main_row = row
                for data_arr in main_mut_dict[cell_id]:
                    main_row += 1
                    tmp_data = [main_homo_hetero, data_arr[0], self.indel_posit, data_arr[7]]
                    self.make_row(sheet, main_row, tmp_data, 4)
                sub_row = row
                if cell_id in sub_mut_dict:
                    for data_arr in sub_mut_dict[cell_id]:
                        sub_row += 1
                        tmp_data = [sub_homo_hetero, data_arr[0], self.indel_posit, data_arr[7]]
                        self.make_row(sheet, sub_row, tmp_data, 8)
                if cell_id in sub_non_mut_dict:
                    for data_arr in sub_non_mut_dict[cell_id]:
                        sub_row += 1
                        tmp_data = [sub_homo_hetero, data_arr[0], self.indel_nega, data_arr[7]]
                        self.make_row(sheet, sub_row, tmp_data, 8)
                tmp_row = row + 1
                row = main_row
                if main_row < sub_row:
                    row = sub_row
                for idx_row in range(tmp_row, row + 1):
                    tmp_data = [str(index), cell_id.split(self.deli)[0], cell_id.split(self.deli)[1]]
                    self.make_row(sheet, idx_row, tmp_data)
                    index += 1

            if cell_id in main_non_mut_dict:
                main_row = row
                for data_arr in main_non_mut_dict[cell_id]:
                    main_row += 1
                    tmp_data = [main_homo_hetero, data_arr[0], self.indel_nega, data_arr[7]]
                    self.make_row(sheet, main_row, tmp_data, 4)
                sub_row = row
                if cell_id in sub_mut_dict:
                    for data_arr in sub_mut_dict[cell_id]:
                        sub_row += 1
                        tmp_data = [sub_homo_hetero, data_arr[0], self.indel_posit, data_arr[7]]
                        self.make_row(sheet, sub_row, tmp_data, 8)
                if cell_id in sub_non_mut_dict:
                    for data_arr in sub_non_mut_dict[cell_id]:
                        sub_row += 1
                        tmp_data = [sub_homo_hetero, data_arr[0], self.indel_nega, data_arr[7]]
                        self.make_row(sheet, sub_row, tmp_data, 8)
                tmp_row = row + 1
                row = main_row
                if main_row < sub_row:
                    row = sub_row
                for idx_row in range(tmp_row, row + 1):
                    tmp_data = [str(index), cell_id.split(self.deli)[0], cell_id.split(self.deli)[1]]
                    self.make_row(sheet, idx_row, tmp_data)
                    index += 1
            # blank row by cell_id
            row += 1

        # make homo_homo, homo_hetero, homo_wt, hetero_homo,... count column
        self.make_row(sheet, 2, cnt_homo_hetero_wt)

        workbook.save(filename=path + self.ext_xlsx)

        return [cell_non_junk_list, non_cell_junk_list]

    def make_excel_by_arr_list(self, path, data_list):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        row = 1
        header_arr = ['barcode_1', 'barcode_2', 'indel_flag', 'Aligned_Sequence', 'Reference_Sequence', 'Read_Status',
                      'n_deleted', 'n_inserted', 'n_mutated', '#Reads', '%Reads']
        self.make_row(sheet, row, header_arr)

        for data_arr in data_list:
            barcode_pair = data_arr[0].split(self.deli)
            mut_list = data_arr[1]
            non_mut_list = data_arr[2]

            for tmp_arr in mut_list:
                row += 1
                tmp_data = [barcode_pair[0], barcode_pair[1], self.indel_posit]
                self.make_row(sheet, row, tmp_data)
                col = 4
                for tmp_data in tmp_arr:
                    sheet.cell(row=row, column=col, value=tmp_data)
                    col += 1

            for tmp_arr in non_mut_list:
                row += 1
                tmp_data = [barcode_pair[0], barcode_pair[1], self.indel_nega]
                self.make_row(sheet, row, tmp_data)
                col = 4
                for tmp_data in tmp_arr:
                    sheet.cell(row=row, column=col, value=tmp_data)
                    col += 1

        workbook.save(filename=path + self.ext_xlsx)

    def make_excel_hom_hete_filter_out_by_frequency(self, path, trgt_list, id_list, thres_arr):
        logic = Logic.Logics()

        workbook = openpyxl.Workbook()
        sheet = workbook.active

        main_mut_dict = trgt_list[0][0]
        main_non_mut_dict = trgt_list[0][1]
        sub_mut_dict = trgt_list[1][0]
        sub_non_mut_dict = trgt_list[1][1]

        threshold_main_freq = thres_arr[0]
        threshold_sub_freq = thres_arr[1]

        row = 1
        cnt_sam = ['homo', 'hetero', 'wt']
        cnt_arr = []
        for h_1 in cnt_sam:
            for h_2 in cnt_sam:
                cnt_arr.append(h_1 + "_" + h_2)
        self.make_row(sheet, row, cnt_arr)

        row = 4
        header_arr = ['index', 'barcode_1', 'barcode_2', 'homo_hetero', 'main_NGS_read', 'indel_flag', '#of_read',
                      'sub_homo_hetero', 'sub_NGS_read', 'sub_indel_flag', 'sub_#of_read']
        self.make_row(sheet, row, header_arr)

        cell_non_junk_list = []
        non_cell_junk_list = []
        cnt_homo_hetero_wt = [0, 0, 0, 0, 0, 0, 0, 0, 0]
        for cell_id in id_list:
            index = 1

            # filter out cell_none or none_cell junks
            main_homo_hetero = logic.check_homo_hetero(cell_id, main_mut_dict, main_non_mut_dict)
            sub_homo_hetero = logic.check_homo_hetero(cell_id, sub_mut_dict, sub_non_mut_dict)
            if main_homo_hetero == "":
                if sub_homo_hetero == "":
                    continue
                else:  # filter out cell_none junks
                    logic.add_junk_list(cell_id, sub_mut_dict, sub_non_mut_dict, non_cell_junk_list)
                    continue
            # filter out none_cell junks
            if sub_homo_hetero == "":
                logic.add_junk_list(cell_id, main_mut_dict, main_non_mut_dict, cell_non_junk_list)
                continue

            # filter out by threshold
            len_main_mut, len_main_non = logic.count_freq_by_cell(cell_id, main_mut_dict, main_non_mut_dict)
            len_sub_mut, len_sub_non = logic.count_freq_by_cell(cell_id, sub_mut_dict, sub_non_mut_dict)
            if len_main_mut + len_main_non <= threshold_main_freq:
                continue
            if len_sub_mut + len_sub_non <= threshold_sub_freq:
                continue

            logic.count_homo_hetero_wt(main_homo_hetero, sub_homo_hetero, cnt_homo_hetero_wt)

            if cell_id in main_mut_dict:
                main_row = row
                for data_arr in main_mut_dict[cell_id]:
                    main_row += 1
                    tmp_data = [main_homo_hetero, data_arr[0], self.indel_posit, data_arr[7]]
                    self.make_row(sheet, main_row, tmp_data, 4)
                sub_row = row
                if cell_id in sub_mut_dict:
                    for data_arr in sub_mut_dict[cell_id]:
                        sub_row += 1
                        tmp_data = [sub_homo_hetero, data_arr[0], self.indel_posit, data_arr[7]]
                        self.make_row(sheet, sub_row, tmp_data, 8)
                if cell_id in sub_non_mut_dict:
                    for data_arr in sub_non_mut_dict[cell_id]:
                        sub_row += 1
                        tmp_data = [sub_homo_hetero, data_arr[0], self.indel_nega, data_arr[7]]
                        self.make_row(sheet, sub_row, tmp_data, 8)
                tmp_row = row + 1
                row = main_row
                if main_row < sub_row:
                    row = sub_row
                for idx_row in range(tmp_row, row + 1):
                    tmp_data = [str(index), cell_id.split(self.deli)[0], cell_id.split(self.deli)[1]]
                    self.make_row(sheet, idx_row, tmp_data)
                    index += 1

            if cell_id in main_non_mut_dict:
                main_row = row
                for data_arr in main_non_mut_dict[cell_id]:
                    main_row += 1
                    tmp_data = [main_homo_hetero, data_arr[0], self.indel_nega, data_arr[7]]
                    self.make_row(sheet, main_row, tmp_data, 4)
                sub_row = row
                if cell_id in sub_mut_dict:
                    for data_arr in sub_mut_dict[cell_id]:
                        sub_row += 1
                        tmp_data = [sub_homo_hetero, data_arr[0], self.indel_posit, data_arr[7]]
                        self.make_row(sheet, sub_row, tmp_data, 8)
                if cell_id in sub_non_mut_dict:
                    for data_arr in sub_non_mut_dict[cell_id]:
                        sub_row += 1
                        tmp_data = [sub_homo_hetero, data_arr[0], self.indel_nega, data_arr[7]]
                        self.make_row(sheet, sub_row, tmp_data, 8)
                tmp_row = row + 1
                row = main_row
                if main_row < sub_row:
                    row = sub_row
                for idx_row in range(tmp_row, row + 1):
                    tmp_data = [str(index), cell_id.split(self.deli)[0], cell_id.split(self.deli)[1]]
                    self.make_row(sheet, idx_row, tmp_data)
                    index += 1
            # blank row by cell_id
            row += 1

        # make homo_homo, homo_hetero, homo_wt, hetero_homo,... count column
        self.make_row(sheet, 2, cnt_homo_hetero_wt)

        workbook.save(filename=path + self.ext_xlsx)

        return [cell_non_junk_list, non_cell_junk_list]

    def make_row(self, sheet, row, data_arr, col=1):
        for idx in range(len(data_arr)):
            sheet.cell(row=row, column=(col + idx), value=data_arr[idx])

    def make_excel_tot_read_by_cell_filter_out_by_frequency(self, path, trgt_list, id_list, thres_arr):
        logic = Logic.Logics()

        workbook = openpyxl.Workbook()
        sheet = workbook.active

        main_mut_dict = trgt_list[0][0]
        main_non_mut_dict = trgt_list[0][1]
        sub_mut_dict = trgt_list[1][0]
        sub_non_mut_dict = trgt_list[1][1]

        threshold_main_freq = thres_arr[0]
        threshold_sub_freq = thres_arr[1]

        row = 1
        cnt_sam = ['homo', 'hetero', 'wt']
        cnt_arr = []
        for h_1 in cnt_sam:
            for h_2 in cnt_sam:
                cnt_arr.append(h_1 + "_" + h_2)
        self.make_row(sheet, row, cnt_arr)

        row = 4
        header_arr = ['index', 'barcode_1', 'barcode_2', 'homo_hetero', '#0f_O', '%of_O', '#of_X', '%of_X',
                      'total_main', 'target_homo_hetero', '#0f_target_O', '%of_target_O', '#of_target_X',
                      '%of_target_X', 'total_target']
        self.make_row(sheet, row, header_arr)

        cell_non_junk_list = []
        non_cell_junk_list = []
        cnt_homo_hetero_wt = [0, 0, 0, 0, 0, 0, 0, 0, 0]
        for cell_id in id_list:

            # filter out cell_none or none_cell junks
            main_homo_hetero = logic.check_homo_hetero(cell_id, main_mut_dict, main_non_mut_dict)
            sub_homo_hetero = logic.check_homo_hetero(cell_id, sub_mut_dict, sub_non_mut_dict)
            if main_homo_hetero == "":
                if sub_homo_hetero == "":
                    continue
                else:  # filter out cell_none junks
                    logic.add_junk_list(cell_id, sub_mut_dict, sub_non_mut_dict, non_cell_junk_list)
                    continue
            # filter out none_cell junks
            if sub_homo_hetero == "":
                logic.add_junk_list(cell_id, main_mut_dict, main_non_mut_dict, cell_non_junk_list)
                continue

            # filter out by threshold
            len_main_mut, len_main_non = logic.count_freq_by_cell(cell_id, main_mut_dict, main_non_mut_dict)
            len_sub_mut, len_sub_non = logic.count_freq_by_cell(cell_id, sub_mut_dict, sub_non_mut_dict)
            if len_main_mut + len_main_non <= threshold_main_freq:
                continue
            if len_sub_mut + len_sub_non <= threshold_sub_freq:
                continue

            logic.count_homo_hetero_wt(main_homo_hetero, sub_homo_hetero, cnt_homo_hetero_wt)

            row += 1
            barcd1 = cell_id.split(self.deli)[0]
            barcd2 = cell_id.split(self.deli)[1]
            tmp_row = [row - 4, barcd1, barcd2, main_homo_hetero, len_main_mut,
                       len_main_mut * self.percent / (len_main_mut + len_main_non), len_main_non,
                       len_main_non * self.percent / (len_main_mut + len_main_non), len_main_mut + len_main_non,
                       sub_homo_hetero, len_sub_mut, len_sub_mut * self.percent / (len_sub_mut + len_sub_non),
                       len_sub_non, len_sub_non * self.percent / (len_sub_mut + len_sub_non), len_sub_mut + len_sub_non]
            self.make_row(sheet, row, tmp_row)

        # make homo_homo, homo_hetero, homo_wt, hetero_homo,... count column
        self.make_row(sheet, 2, cnt_homo_hetero_wt)

        workbook.save(filename=path + self.ext_xlsx)

        return [cell_non_junk_list, non_cell_junk_list]

    def make_excel_by_list(self, path, result_list, cnt_homo_hetero_wt):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        row = 1
        cnt_sam = ['homo', 'hetero', 'wt']
        cnt_arr = []
        for h_1 in cnt_sam:
            for h_2 in cnt_sam:
                cnt_arr.append(h_1 + "_" + h_2)
        self.make_row(sheet, row, cnt_arr)

        # make homo_homo, homo_hetero, homo_wt, hetero_homo,... count column
        self.make_row(sheet, 2, cnt_homo_hetero_wt)

        row = 4
        header_arr = ['index', 'barcode_1', 'barcode_2', 'homo_hetero', '#0f_O', '%of_O', '#of_X', '%of_X',
                      'total_main', 'target_homo_hetero', '#0f_target_O', '%of_target_O', '#of_target_X',
                      '%of_target_X', 'total_target']
        self.make_row(sheet, row, header_arr)

        for result_arr in result_list:
            row += 1
            self.make_row(sheet, row, result_arr)

        workbook.save(filename=path + self.ext_xlsx)


